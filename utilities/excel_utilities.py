# import libraries
import openpyxl


class Xlutils:
    # constructor
    def __init__(self, path) -> None:
        self.path = path

    # return count of rows
    def get_row_count(self, sheet_name):
        wb = openpyxl.load_workbook(self.path)
        ws = wb[sheet_name]
        return ws.max_row

    # returns count of columns
    def get_column_count(self, sheet_name):
        wb = openpyxl.load_workbook(self.path)
        ws = wb[sheet_name]
        return ws.max_column

    # returns cell data
    def get_data(self, row, col, sheet_name):
        wb = openpyxl.load_workbook(self.path)
        ws = wb[sheet_name]
        return ws.cell(row, col).value
